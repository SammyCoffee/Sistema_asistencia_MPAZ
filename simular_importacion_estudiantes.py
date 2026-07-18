from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

from base_datos import obtener_conexion
from revisar_excel_estudiantes import (
    limpiar_texto,
    normalizar_rut,
    rut_es_valido
)

RUTA_EXCEL = Path(
    "datos_privados/REGISTRO DATOS.xlsx"

)

RUTA_INFORME = Path(
    "datos_privados/simulacion_importacion_estudiantes.txt"
)

NOMBRE_HOJA = "ESTUDIANTE"


def obtener_ruts_existentes():
    conexion = obtener_conexion()
    cursor = conexion.cursor()


    try:

        cursor.execute(
            """
            SELECT rut
            FROM alumnos
            """

        )

        registros = cursor.fetchall()

    finally:
        conexion.close()

    ruts_existentes = set()

    for registro in registros:
        rut_normalizado = normalizar_rut(
            registro[0]

        )

        if rut_normalizado:
            ruts_existentes.add(
                rut_normalizado
            )        

    return ruts_existentes

def leer_estudiantes_excel():
    libro = load_workbook(
        RUTA_EXCEL,
                read_only=True,
                data_only=True
    )           

    hoja = libro[NOMBRE_HOJA]

    estudiantes = []

    for numero_fila in range(
        3,
        hoja.max_row + 1
    ):
        curso = limpiar_texto(
            hoja.cell(numero_fila, 4).value
        ).upper()

        nombres = limpiar_texto(
            hoja.cell(numero_fila, 6).value
        )

        apellidos = limpiar_texto(
            hoja.cell(numero_fila, 7).value
        )
        rut_original = limpiar_texto(
            hoja.cell(numero_fila, 8).value
        )

        if (
            not curso
            and not nombres
            and not apellidos
            and not rut_original
        ):
            continue

        rut = normalizar_rut(
            rut_original
        )

        estudiantes.append(
            {
                "fila": numero_fila,
                "rut": rut,
                "rut_original": rut_original,
                "nombre_completo": limpiar_texto(
                    f"{nombres} {apellidos}"
                ),
                "curso": curso
            }
        )
    libro.close()

    return estudiantes

def simular_importacion():
    if not RUTA_EXCEL.exists():
        print("No se encontro el Excel:" )
        print(RUTA_EXCEL)
        return

    estudiantes = leer_estudiantes_excel()

    ruts_excel = Counter(
        estudiante["rut"]
        for estudiante in estudiantes
        if estudiante["rut"]
    )


    ruts_existentes = obtener_ruts_existentes()

    estudiantes_aptos = []
    estudiantes_omitidos = []

    for estudiante in estudiantes:
        motivos = []

        rut = estudiante["rut"]
        nombre = estudiante["nombre_completo"]
        curso = estudiante["curso"]


        if not rut:
            motivos.append(
                "RUT vacio o con formato incorrecto"
            )
        else:
            if not rut_es_valido(rut):
                motivos.append(
                    "RUT con digito verificado invalido"
                )        
            if ruts_excel[rut] > 1:
                motivos.append(
                    "RUT repetido dentro del Excel"
                )
            if rut in ruts_existentes:
                motivos.append(
                "El alumno ya existe en SQLite"
                )        

        if not nombre:
            motivos.append(
                "Nombre incompletos"
            )    
        if not curso:
            motivos.append(
                "Curso vacio"
            )    
        if motivos:
            estudiantes_omitidos.append(
                {
                    **estudiante,
                    "motivos": motivos
                }
            )
        else:
            estudiantes_aptos.append(
                estudiante
            )       
    lineas = []

    lineas.append(
            "SIMULACION DE IMPORTANCIA DE ESTUDIANTES"
        )    
    lineas.append("=" * 50)
    lineas.append(
            f"Registros leidos del Excel: {len(estudiantes)}"    
        )
    lineas.append(
            f"Aptos para importar: {len(estudiantes_aptos)}"
        )
    lineas.append(
            f"Omitidos para revision: {len(estudiantes_omitidos)}"
        )
    lineas.append(
            f"Alumnos actuales en SQLite: {len(ruts_existentes)}"
        )

    lineas.append("")
    lineas.append(
            "REGISTRO OMITIDOS"
        )
    lineas.append("-" * 50)

    if estudiantes_omitidos:
        for estudiante in estudiantes_omitidos:
            motivos = ";".join(
                estudiante["motivos"]
            )

            rut_mostrado = (
                estudiante["rut"]
                or estudiante["rut_original"]
                or "Sin RUT"
            )

            lineas.append(
                f"Fila {estudiante['fila']} | "
                f"RUT: {rut_mostrado} | "
                f"Curso: {estudiante['curso']} | "
                f"Motivo: {motivos}"
            )
    else:
        lineas.append(
            "No existen registros omnitidos"
        )

    lineas.append("")
    lineas.append(
        "RESUMEN POR CURSO DE LOS APTOS"
    )    
    lineas.append("-" * 50)

    cantidad_por_curso = Counter(
        estudiante["curso"]
        for estudiante in estudiantes_aptos
    )

    for curso in sorted(cantidad_por_curso):
        lineas.append(
            f"{curso}: {cantidad_por_curso[curso]}"
        )
    RUTA_INFORME.parent.mkdir(
        parents=True,
        exist_ok=True
    )    

    RUTA_INFORME.write_text(
                "\n".join(lineas),
                encoding="utf-8"
    )    

    print("Simulacion completada correctamente")
    print(
        "Registro leidos:",
        len(estudiantes)
    )
    print(
            "Aptos apra importar:",
            len(estudiantes_aptos)
    )
    print(
        "Omitidos para revision:",
        len(estudiantes_omitidos)
    )
    print(
        "La base de datos No fue modificada"
    )
    print(
        "Informe privado:",
        RUTA_INFORME
    )

if __name__ == "__main__":
    simular_importacion()        