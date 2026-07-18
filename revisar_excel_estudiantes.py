from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


RUTA_EXCEL = Path(
    "datos_privados/REGISTRO DATOS.xlsx"
)

RUTA_INFORME = Path(
    "datos_privados/revision_estudiantes.txt"
)

NOMBRE_HOJA = "ESTUDIANTE"

def limpiar_texto(valor):
    if valor is None:
        return ""
    
    return " ".join(
        str(valor).strip().split()
    )

def normalizar_rut(rut):
    rut = limpiar_texto(rut).upper()

    rut = (
        rut.replace(".", "")
        .replace("-", "")
        .replace(" ", "")
    )

    if len(rut) < 2:
        return ""
    
    cuerpo = rut[:-1]
    digito_verificador = rut[-1]


    if not cuerpo.isdigit():
        return ""
    
    if not (
        digito_verificador.isdigit()
        or digito_verificador == "K"
    ):
        return ""
    
    return f"{cuerpo}-{digito_verificador}"

def rut_es_valido(rut):
    rut_normalizado = normalizar_rut(rut)

    if not rut_normalizado:
        return False
    
    cuerpo, digito_ingresado = rut_normalizado.split("-")

    suma = 0

    multiplicador = 2

    for numero in reversed(cuerpo):
        suma += int(numero) * multiplicador

        multiplicador += 1

        if multiplicador > 7:
            multiplicador = 2
    
    resultado = 11 - (suma % 11)

    if resultado == 11:
        digito_calculado = "0"

    elif resultado == 10:
        digito_calculado = "K"

    else: 
        digito_calculado = str(resultado)


    return digito_ingresado == digito_calculado

def revisar_excel():
    if not RUTA_EXCEL.exists():
        print("No se encontro el archivo:")
        print(RUTA_EXCEL)
        return

    libro = load_workbook(
        RUTA_EXCEL,
        read_only=True,
        data_only=True
    )

    if NOMBRE_HOJA not in libro.sheetnames:
        print(
            "No existe la hoja:",
            NOMBRE_HOJA
        )
        libro.close()
        return
    hoja = libro[NOMBRE_HOJA]

    estudiantes = []
    campos_incompletos = []
    ruts_invalidos = []


    for numero_fila in range(3, hoja.max_row +1):
        curso = limpiar_texto(
            hoja.cell(numero_fila, 4).value
        ).upper()

        nombres = limpiar_texto(
            hoja.cell(numero_fila, 6).value
        )              

        apellidos = limpiar_texto(
            hoja.cell(numero_fila, 7).value
        )

        rut_original = limpiar_texto(
            hoja.cell(numero_fila, 8).value
        )

        if (
            not curso 
            and not nombres 
            and not apellidos 
            and not rut_original
        ):    
            continue

        rut = normalizar_rut(rut_original)

        estudiante = {
            "fila": numero_fila,
            "rut": rut,
            "rut_original": rut_original,
            "nombre_completo": limpiar_texto(
                f"{nombres} {apellidos}"
            ),
            "curso": curso
        }

        estudiantes.append(estudiante)

        campos_faltantes = []

        if not rut:
            campos_faltantes.append("RUT")

        if not nombres:
            campos_faltantes.append("nombres")

        if not apellidos:
            campos_faltantes.append("apellidos")

        if not curso:
            campos_faltantes.append("curso")            

        if campos_faltantes:
            campos_incompletos.append(
                {
                    "fila": numero_fila,
                    "campos": campos_faltantes
                }
            )

        if rut and not rut_es_valido(rut):
            ruts_invalidos.append(
                {
                    "fila": numero_fila,
                    "rut": rut
                }
            )    
    libro.close()

    filas_por_rut = defaultdict(list)

    for estudiante in estudiantes:
        if estudiante["rut"]:
            filas_por_rut[
                estudiante["rut"]
            ].append(estudiante["fila"])

    ruts_repetidos = {
        rut: filas
        for rut, filas in filas_por_rut.items()
        if len(filas) > 1
    }        

    cantidad_por_curso = Counter(
        estudiante["curso"]
        for estudiante in estudiantes
        if estudiante["curso"]
    )

    lineas = []

    lineas.append("REVISION DEL EXCEL DE ESTUDIANTES")
    lineas.append("=" * 40)
    lineas.append(
        F"Estudiantes encontrados: {len(estudiantes)}"
    )
    lineas.append(
        f"RUT repetidos: {len(ruts_repetidos)}"
    )
    lineas.append(
        f"RUT posiblemente invalidos: {len(ruts_invalidos)}"
    )
    lineas.append(

        f"Filas con campos incompletos: "
        f"{len(campos_incompletos)}"
    )

    lineas.append("")
    lineas.append("ESTUDIANTES POR CURSO")
    lineas.append("-" * 40)

    for curso in sorted(cantidad_por_curso):
        cantidad = cantidad_por_curso[curso]

        lineas.append(
            f"{curso}: {cantidad}"
        )

    lineas.append("")
    lineas.append("RUT REPETIDOS")
    lineas.append("-" * 40)

    if ruts_repetidos:
        for rut, filas in ruts_repetidos.items():
            lineas.append(
                f"{rut} aparece en las filas: {filas}"
            )

    else:
        lineas.append("No se encontraton RUT repetidos")

    lineas.append("")
    lineas.append("RUT PARA REVISAR")
    lineas.append("-" * 40)  

    if ruts_invalidos:
        for elemento in ruts_invalidos:
            lineas.append(
                f"Fila {elemento['fila']}: "
                f"{elemento['rut']}"
            )
    else:
        lineas.append("No se encontraron RUT repetidos")          

    lineas.append("")
    lineas.append("CAMPOS INCOMPLETOS")
    lineas.append("-" * 40)            

    if campos_incompletos:
        for elemento in campos_incompletos:
            campos = ", ".join(
                elemento["campos"]
            )

            lineas.append(
                f"Fila {elemento['fila']}: {campos}"
            )
    else:
        lineas.append(
            "No se encontaron campos incompletos"
        )        

    contenido_informe = "\n".join(lineas)

    RUTA_INFORME.write_text(
        contenido_informe,
        encoding="utf-8"
    )    

    print("Revision completa correctamente")
    print("Estudiante encontrados:", len(estudiantes))
    print("RUT repetidos:", len(ruts_repetidos))
    print(
        "RUT para revisar:",
        len(ruts_invalidos)
    )
    print(
        "Filas incompletas:",
        len(campos_incompletos)
    )
    print("Informe privado:", RUTA_INFORME)


if __name__ == "__main__":
    revisar_excel()    